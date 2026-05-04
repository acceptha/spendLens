from pathlib import Path

import openpyxl
import pytest

from app.parsers import ParseError
from app.parsers.samsung_card import REQUIRED_COLUMNS, find_header_row, find_target_sheet

FIXTURE = Path(__file__).resolve().parents[1] / "fixtures" / "samsung-card-fixture.xlsx"


def test_find_header_row_returns_4():
    wb = openpyxl.load_workbook(FIXTURE, read_only=False, data_only=True)
    sheet_name = find_target_sheet(wb)
    ws = wb[sheet_name]
    row_idx, col_map = find_header_row(ws)
    assert row_idx == 4
    for c in REQUIRED_COLUMNS:
        assert c in col_map


def test_find_header_row_raises_when_required_missing():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "엉뚱한 헤더"
    with pytest.raises(ParseError) as ei:
        find_header_row(ws)
    assert ei.value.code == "HEADER_NOT_FOUND"
