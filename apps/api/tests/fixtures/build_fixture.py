"""Build the samsung-card fixture XLSX. Run once to generate fixture file.

Usage:
    cd apps/api && uv run python tests/fixtures/build_fixture.py
"""
from pathlib import Path

import openpyxl


def build() -> None:
    wb = openpyxl.Workbook()
    other = wb.active
    other.title = "전체이용내역"
    other["A1"] = "(unused for W1)"

    ws = wb.create_sheet("■ 국내이용내역")
    # 1~3행: 헤더 위 메타 (실제 명세서 모방)
    ws["A1"] = "삼성카드 국내이용내역"
    ws["A2"] = "조회기간: 2026-04-01 ~ 2026-04-30"
    # 4행: 헤더
    headers = [
        "카드번호", "본인가족구분", "승인일자", "승인시각", "가맹점명",
        "승인금액(원)", "일시불할부구분", "할부개월", "승인번호",
        "취소여부", "사용포인트", "결제일",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col_idx, value=h)

    # 5행~: 데이터
    rows = [
        ["1234-5678-9012-3456", "본인", "2026-04-28", "12:34:00", "스타벅스 강남대로점",
         9500, "일시불", 0, "A20260428001", "N", 0, "2026-05-25"],
        ["1234-5678-9012-3456", "본인", "2026-04-28", "19:42:00", "이태원 BBQ",
         18200, "일시불", 0, "A20260428002", "N", 0, "2026-05-25"],
        ["1234-5678-9012-3456", "본인", "2026-04-27", "12:11:00", "김밥천국 역삼점",
         11000, "일시불", 0, "A20260427001", "N", 0, "2026-05-25"],
        ["1234-5678-9012-3456", "본인", "2026-04-26", None, "이마트 잠실점",
         59800, "할부", 3, "A20260426001", "N", 0, "2026-05-25"],
        ["1234-5678-9012-3456", "본인", "2026-04-25", "13:22:00", "이디야커피 선릉점",
         5500, "일시불", 0, "A20260425001", "N", 0, "2026-05-25"],
        ["1234-5678-9012-3456", "본인", "2026-04-24", "20:00:00", "교보문고 강남점",
         24000, "일시불", 0, "A20260424001", "Y", 0, "2026-05-25"],  # 취소
        ["1234-5678-9012-3456", "본인", "2026-04-23", "00:00:00", "정산수수료",
         500, "일시불", 0, "", "N", 0, "2026-05-25"],  # 승인번호 누락
    ]
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    out = Path(__file__).parent / "samsung-card-fixture.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    build()
