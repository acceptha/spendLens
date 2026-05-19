"""Build the Woori card fixture XLSX. Run once.

Usage:
    cd apps/api && uv run python tests/fixtures/build_woori_fixture.py
"""
from pathlib import Path

import openpyxl


def build() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sheet 1"

    # Row 1: title
    ws["A1"] = "   이용대금명세서 상세 내역"

    # Row 2: primary header (with embedded \n as in real statement)
    headers_r2 = [
        "이용\n일자", "카드\n구분", "이용\n카드", "매출\n구분",
        "이용가맹점(은행)명", "이용금액\n(해외현지/\n체크카드)",
        "할부\n개월", "당월결제하실금액",
        None, None, None, None,
        "결제후 잔액", "할부가격",
    ]
    for c_idx, h in enumerate(headers_r2, start=1):
        ws.cell(row=2, column=c_idx, value=h)

    # Row 3: sub-header for cols 8-12
    sub_headers_r3 = [None, None, None, None, None, None, None,
                     "회차", "원금", "혜택금액", "환율", "수수료", None, None]
    for c_idx, h in enumerate(sub_headers_r3, start=1):
        if h:
            ws.cell(row=3, column=c_idx, value=h)

    # Data rows from row 4 onward — fictitious merchants matching the real statement layout
    rows = [
        ["04.14", "체크/본인", "5102", "국내체크", "쿠팡-쿠팡(주)", "130,900", "0",
         "0", "0", "0", "0", "0", "0", "0"],
        ["04.02", "신용/본인", "6247", "국내일시불", "씨유(CU) 구로JNK점", "1,800", "0",
         "0", "0", "1,777", "23", "0", "0", "0"],
        ["04.03", "신용/본인", "6247", "국내일시불", "스타벅스 강남점", "5,400", "0",
         "0", "0", "5,335", "65", "0", "0", "0"],
        ["04.06", "신용/본인", "6247", "국내일시불", "홈플러스", "47,630", "0",
         "0", "0", "47,249", "381", "0", "0", "0"],
        ["04.08", "신용/본인", "6247", "할부3개월", "이마트 성수점", "60,000", "3",
         "0", "0", "60,000", "0", "0", "0", "0"],
        ["04.10", "신용/본인", "6247", "국내일시불", "넷플릭스", "13,500", "0",
         "0", "0", "13,392", "108", "0", "0", "0"],
        ["04.18", "신용/본인", "6247", "국내체크", "지에스25 시흥5동점", "800", "0",
         "0", "0", "0", "0", "0", "0", "0"],
        ["04.18", "신용/본인", "6247", "취소", "취소-지에스25 시흥5동점", "-800", "0",
         "0", "0", "0", "0", "0", "0", "0"],
        ["05.01", "신용/본인", "6247", "국내일시불", "교통-지하철 8건", "12,950", "0",
         "0", "0", "12,782", "168", "0", "0", "0"],
        ["05.01", "신용/본인", "6247", "국내일시불", "올댓폰", "8,000", "0",
         "0", "0", "7,936", "64", "0", "0", "0"],
        # Summary rows (must be skipped by parser)
        [None, "신용/본인", "6247", "카드소계", "소계(하수임)", None, "0",
         "0", "0", "150,471", "809", "0", "0", "0"],
        [None, "신용/본인", "6247", "통합청구합계", "통합청구합계", None, "0",
         "0", "0", "150,471", "809", "0", "0", "0"],
    ]
    for r_offset, row in enumerate(rows, start=4):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_offset, column=c_idx, value=val)

    out = Path(__file__).parent / "woori-card-fixture.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    build()
