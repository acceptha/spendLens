"""Build the 하나은행 fixture XLSX. Run once.

Usage:
    cd apps/api && uv run python tests/fixtures/build_hana_bank_fixture.py
"""
from pathlib import Path

import openpyxl


def build() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Row 1: title
    ws["A1"] = "거래내역조회"

    # Row 3: account meta (informational only)
    ws["A3"] = "계좌번호"
    ws["B3"] = "111-222333-44444"
    ws["D3"] = "예금종류"
    ws["E3"] = "수신"
    ws["G3"] = "조회기간"
    ws["H3"] = "2026-04-01 ~ 2026-04-30"

    # Row 4: balance summary
    ws["A4"] = "잔액"
    ws["B4"] = 1234567
    ws["D4"] = "인출가능금액"
    ws["E4"] = 1234567

    # Row 6: data header
    headers = ["거래일시", "구분", "적요", "출금액", "입금액", "잔액", "거래점"]
    for c_idx, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c_idx, value=h)

    # Row 7+: data rows (fictitious)
    rows = [
        ("2026-04-28 12:34:56", "타행이체", "월급-주식회사 가상", 0, 3500000, 4734567, "우리은행(0001)"),  # noqa: E501
        ("2026-04-28 19:42:00", "현금/체크", "이마트 잠실", 47000, 0, 1234567, "광안동"),
        ("2026-04-27 12:11:00", "CMS", "넷플릭스 정기결제", 13500, 0, 1281567, "광안동"),
        ("2026-04-26 09:00:00", "타행이체", "엄마", 0, 200000, 1295067, "신한은행(2345)"),
        ("2026-04-25 13:22:00", "현금/체크", "스타벅스 강남", 5500, 0, 1095067, "광안동"),
        ("2026-04-24 20:00:00", "정기적금", "청년도약계좌", 300000, 0, 1100567, "광안동"),
        ("2026-04-23 00:00:00", "CMS", "월세-임대인", 800000, 0, 1400567, "광안동"),
        ("2026-04-22 14:00:00", "타행이체", "친구1", 0, 50000, 2200567, "국민은행(8888)"),
        ("2026-04-21 08:15:00", "현금/체크", "GS25 강남대로점", 3200, 0, 2150567, "광안동"),
        ("2026-04-20 19:00:00", "현금/체크", "BBQ 강남", 24000, 0, 2153767, "광안동"),
    ]
    for r_offset, row in enumerate(rows, start=7):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_offset, column=c_idx, value=val)

    out = Path(__file__).parent / "hana-bank-fixture.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    build()
