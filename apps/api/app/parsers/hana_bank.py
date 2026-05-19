"""하나은행 통장 거래내역 XLSX 파서.

Sheet: 'Sheet1'
Row 1: 제목 '거래내역조회'
Row 6: 데이터 헤더 ('거래일시', '구분', '적요', '출금액', '입금액', '잔액', '거래점')
Row 7+: 데이터

통장이라 카드 정보(승인번호/last4/할부) 없음.
amount: 출금=양수, 입금=음수 (sign으로 흐름 구분).
merchant_raw: '[구분] 적요' 형태로 결합.
"""
import io
from datetime import datetime
from decimal import Decimal
from typing import Any

import openpyxl

from app.parsers import ParseError
from app.parsers.samsung_card import ParseResult

_TITLE_KEYWORD = "거래내역조회"
REQUIRED_HEADERS = ["거래일시", "출금액", "입금액", "적요"]


def detect(wb: openpyxl.Workbook) -> bool:
    """Return True if any sheet has row1 title '거래내역조회'
    AND a header row containing all REQUIRED_HEADERS.
    """
    for name in wb.sheetnames:
        ws = wb[name]
        # Row 1 title check via iter_rows (read_only-safe)
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            continue
        a1 = first_row[0]
        if not a1 or _TITLE_KEYWORD not in str(a1):
            continue
        # Scan rows 1-10 for required headers
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            cells = {str(c).strip() for c in row if c is not None}
            if all(h in cells for h in REQUIRED_HEADERS):
                return True
    return False


def _find_target_sheet(wb: openpyxl.Workbook) -> str:
    for name in wb.sheetnames:
        a1 = wb[name].cell(row=1, column=1).value
        if a1 and _TITLE_KEYWORD in str(a1):
            return name
    raise ParseError("SHEET_NOT_FOUND", looking_for=_TITLE_KEYWORD, found=list(wb.sheetnames))


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    """Scan up to row 10 for the data header row. Return (row_idx, {header: col_idx_1based})."""
    for row_idx in range(1, 11):
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            val = cell.value
            if isinstance(val, str):
                col_map[val.strip()] = col_idx
        if all(h in col_map for h in REQUIRED_HEADERS):
            return row_idx, col_map
    raise ParseError("HEADER_NOT_FOUND", required=REQUIRED_HEADERS, scanned_rows=10)


def _parse_datetime(v: Any) -> datetime:
    """Parse 'YYYY-MM-DD HH:MM:SS' or accept datetime directly."""
    if isinstance(v, datetime):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    raise ValueError(f"unparseable datetime: {v!r}")


def _to_int_amount(v: Any) -> int:
    """Parse withdrawal/deposit amount cell to int. 0 if empty/None."""
    if v in (None, ""):
        return 0
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    if isinstance(v, str):
        s = v.replace(",", "").strip()
        if not s:
            return 0
        try:
            return int(s)
        except ValueError:
            try:
                return int(float(s))
            except ValueError:
                return 0
    return 0


def parse_workbook(file_bytes: bytes) -> ParseResult:
    from app.transactions.schemas import TransactionIn

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=False, data_only=True
        )
    except Exception as e:
        raise ParseError("WORKBOOK_LOAD_FAILED", reason=str(e)) from e

    sheet_name = _find_target_sheet(wb)
    ws = wb[sheet_name]
    header_row, col_map = _find_header_row(ws)

    dt_col = col_map["거래일시"]
    type_col = col_map.get("구분")
    memo_col = col_map.get("적요")
    out_col = col_map["출금액"]
    in_col = col_map["입금액"]

    transactions: list[TransactionIn] = []
    errors: list[dict[str, Any]] = []
    data_rows = 0

    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        dt_val = ws.cell(row=row_idx, column=dt_col).value
        if dt_val is None or (isinstance(dt_val, str) and not dt_val.strip()):
            continue

        data_rows += 1

        try:
            dt = _parse_datetime(dt_val)
            out_amt = _to_int_amount(ws.cell(row=row_idx, column=out_col).value)
            in_amt = _to_int_amount(ws.cell(row=row_idx, column=in_col).value)

            # 출금 양수, 입금 음수. 둘 다 0이면 skip (의미 없는 행).
            if out_amt == 0 and in_amt == 0:
                continue
            amount = Decimal(out_amt) if out_amt > 0 else Decimal(-in_amt)

            type_val = ws.cell(row=row_idx, column=type_col).value if type_col else None
            memo_val = ws.cell(row=row_idx, column=memo_col).value if memo_col else None
            type_str = str(type_val).strip() if type_val else ""
            memo_str = str(memo_val).strip() if memo_val else ""

            merchant = f"[{type_str}] {memo_str}".strip() if type_str else memo_str or "(no memo)"

            raw_row = {}
            for header, c_idx in col_map.items():
                raw_row[header] = ws.cell(row=row_idx, column=c_idx).value

            transactions.append(TransactionIn(
                txn_date=dt.date(),
                txn_time=dt.time(),
                amount=amount,
                merchant_raw=merchant,
                approval_no=None,
                card_last4=None,
                installment_months=None,
                is_canceled=False,
                raw_row=raw_row,
            ))
        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})

    if data_rows == 0:
        raise ParseError("EMPTY_SHEET", sheet=sheet_name)

    return ParseResult(rows_total=data_rows, transactions=transactions, parse_errors=errors)
