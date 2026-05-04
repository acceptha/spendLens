import io
import re
from dataclasses import dataclass, field
from datetime import date as _date
from datetime import datetime, time
from decimal import Decimal
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from app.transactions.schemas import TransactionIn

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.parsers import ParseError

_TARGET_SHEET_KEYWORD = "국내이용내역"

REQUIRED_COLUMNS = ["승인일자", "가맹점명", "승인금액(원)", "승인번호", "카드번호"]
ALL_KNOWN_COLUMNS = [
    "카드번호", "본인가족구분", "승인일자", "승인시각", "가맹점명",
    "승인금액(원)", "일시불할부구분", "할부개월", "승인번호",
    "취소여부", "사용포인트", "결제일",
]


def find_target_sheet(wb: openpyxl.Workbook) -> str:
    """Return first sheet name containing '국내이용내역'."""
    for name in wb.sheetnames:
        if _TARGET_SHEET_KEYWORD in name:
            return name
    raise ParseError(
        "SHEET_NOT_FOUND", looking_for=_TARGET_SHEET_KEYWORD, found=list(wb.sheetnames)
    )


def find_header_row(ws: Worksheet) -> tuple[int, dict[str, int]]:
    """Scan rows top-down; return (row_index, {column_name: column_index_1based}).

    A row qualifies if it contains all REQUIRED_COLUMNS as cell values.
    """
    max_scan = min(ws.max_row or 20, 20)
    for row_idx in range(1, max_scan + 1):
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            if isinstance(cell.value, str):
                col_map[cell.value.strip()] = col_idx
        if all(req in col_map for req in REQUIRED_COLUMNS):
            return row_idx, col_map

    raise ParseError(
        "HEADER_NOT_FOUND",
        required=REQUIRED_COLUMNS,
        scanned_rows=max_scan,
    )


def mask_pan(pan: str | None) -> tuple[str, str]:
    """Return (masked_string, last4)."""
    if not pan:
        return "****-****-****-****", ""
    digits = re.sub(r"\D", "", pan)
    last4 = digits[-4:] if len(digits) >= 4 else ""
    masked = f"****-****-****-{last4}" if last4 else "****-****-****-****"
    return masked, last4


def _to_date(v: Any) -> _date:
    if isinstance(v, _date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        return datetime.strptime(v, "%Y-%m-%d").date()
    raise ValueError(f"unparseable date: {v!r}")


def _to_time(v: Any) -> time | None:
    if v is None or v == "":
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, str):
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(v, fmt).time()
            except ValueError:
                continue
        return None
    return None


def _to_decimal(v: Any) -> Decimal:
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    if isinstance(v, str):
        cleaned = v.replace(",", "").strip()
        return Decimal(cleaned)
    raise ValueError(f"unparseable amount: {v!r}")


def _to_int(v: Any) -> int | None:
    if v in (None, ""):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return int(s)
        except ValueError:
            return None
    return None


def parse_row(row: dict[str, Any]) -> "TransactionIn":
    from app.parsers.simple_rules import classify
    from app.transactions.schemas import TransactionIn

    pan = row.get("카드번호") or ""
    masked, last4 = mask_pan(str(pan))

    raw_row = dict(row)
    raw_row["카드번호"] = masked

    approval_raw = row.get("승인번호")
    approval_no: str | None = None
    if approval_raw is not None:
        s = str(approval_raw).strip()
        approval_no = s if s else None

    merchant = str(row.get("가맹점명") or "").strip()
    is_canceled = str(row.get("취소여부") or "").strip().upper() == "Y"

    return TransactionIn(
        txn_date=_to_date(row.get("승인일자")),
        txn_time=_to_time(row.get("승인시각")),
        amount=_to_decimal(row.get("승인금액(원)")),
        merchant_raw=merchant,
        approval_no=approval_no,
        card_last4=last4 or None,
        installment_months=_to_int(row.get("할부개월")),
        is_canceled=is_canceled,
        category=classify(merchant),
        raw_row=raw_row,
    )


@dataclass
class ParseResult:
    rows_total: int
    transactions: list = field(default_factory=list)
    parse_errors: list[dict[str, Any]] = field(default_factory=list)


def parse_workbook(file_bytes: bytes) -> ParseResult:
    from app.transactions.schemas import TransactionIn

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=False, data_only=True
        )
    except Exception as e:
        raise ParseError("WORKBOOK_LOAD_FAILED", reason=str(e)) from e

    sheet_name = find_target_sheet(wb)
    ws = wb[sheet_name]
    header_row, col_map = find_header_row(ws)

    transactions: list[TransactionIn] = []
    errors: list[dict[str, Any]] = []
    data_rows = 0

    max_row = ws.max_row or header_row
    for row_idx in range(header_row + 1, max_row + 1):
        row_dict: dict[str, Any] = {}
        for col_name in ALL_KNOWN_COLUMNS:
            col_idx = col_map.get(col_name)
            row_dict[col_name] = ws.cell(row=row_idx, column=col_idx).value if col_idx else None

        # 빈 행/합계 행 skip: 핵심 필드(승인일자 + 가맹점명) 둘 다 없으면 무시
        if not row_dict.get("승인일자") and not row_dict.get("가맹점명"):
            continue

        data_rows += 1
        try:
            transactions.append(parse_row(row_dict))
        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})

    if data_rows == 0:
        raise ParseError("EMPTY_SHEET", sheet=sheet_name)

    return ParseResult(rows_total=data_rows, transactions=transactions, parse_errors=errors)
