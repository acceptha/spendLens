"""우리카드 XLSX 명세서 파서.

Sheet: 'sheet 1' (영문+공백)
Row 1: 제목 "이용대금명세서 상세 내역"
Row 2: 헤더 (multi-line, \n 포함)
Row 3: 일부 sub-header
Row 4+: 데이터

승인번호 컬럼 없음 → dedup은 fallback (date+amount+merchant) 사용.
일자는 MM.DD 포맷 → 연도 추론 (최대 일자가 today 이후면 작년).
"""
import io
import re
from datetime import date as _date
from datetime import datetime
from decimal import Decimal
from typing import Any

import openpyxl

from app.parsers import ParseError
from app.parsers.samsung_card import ParseResult

_TITLE_KEYWORD = "이용대금명세서"
# Normalized header strings (after stripping \n and whitespace)
REQUIRED_HEADERS = ["이용일자", "이용가맹점(은행)명"]
_SKIP_SALES_TYPES = {
    "카드소계", "청구결제번호소계", "합계", "통합청구합계",
    "입금완료", "소계(하수임)",
}


def _norm_header(s: Any) -> str:
    """Strip newlines and whitespace from header cell value."""
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s))


def detect(wb: openpyxl.Workbook) -> bool:
    """Return True if any sheet has a row1 title containing '이용대금명세서'
    AND a header row containing '이용일자' + '이용가맹점(은행)명'.
    """
    for name in wb.sheetnames:
        ws = wb[name]
        # Check row 1 title
        a1 = ws.cell(row=1, column=1).value
        if not a1 or _TITLE_KEYWORD not in str(a1):
            continue
        # Scan first 5 rows for required headers
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            normalized = {_norm_header(c) for c in row if c is not None}
            if all(h in normalized for h in [_norm_header(r) for r in REQUIRED_HEADERS]):
                return True
    return False


def _find_target_sheet(wb: openpyxl.Workbook) -> str:
    for name in wb.sheetnames:
        a1 = wb[name].cell(row=1, column=1).value
        if a1 and _TITLE_KEYWORD in str(a1):
            return name
    raise ParseError("SHEET_NOT_FOUND", looking_for=_TITLE_KEYWORD, found=list(wb.sheetnames))


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    """Scan up to row 5 for the header. Return (row_idx, {normalized_header: col_idx_1based})."""
    required_norm = [_norm_header(h) for h in REQUIRED_HEADERS]
    for row_idx in range(1, 6):
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            n = _norm_header(cell.value)
            if n:
                col_map[n] = col_idx
        if all(h in col_map for h in required_norm):
            return row_idx, col_map
    raise ParseError("HEADER_NOT_FOUND", required=REQUIRED_HEADERS, scanned_rows=5)


def _parse_md_date(v: Any, year: int) -> _date:
    """Parse 'MM.DD' string into date with the given year. Raise ValueError on failure."""
    if isinstance(v, _date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        s = v.strip()
        m = re.match(r"^(\d{1,2})[.\-/](\d{1,2})$", s)
        if m:
            month, day = int(m.group(1)), int(m.group(2))
            return _date(year, month, day)
    raise ValueError(f"unparseable date: {v!r}")


def _infer_year(raw_dates: list[str]) -> int:
    """Use today's year unless the largest MM.DD is in the future → use last year."""
    today = datetime.now().date()
    max_md: tuple[int, int] = (0, 0)
    for s in raw_dates:
        m = re.match(r"^(\d{1,2})[.\-/](\d{1,2})$", str(s).strip())
        if m:
            md = (int(m.group(1)), int(m.group(2)))
            if md > max_md:
                max_md = md
    if max_md == (0, 0):
        return today.year
    # If max month/day is in the future compared to today → statement crosses year boundary
    candidate = _date(today.year, max_md[0], max_md[1])
    if candidate > today:
        return today.year - 1
    return today.year


def _to_decimal(v: Any) -> Decimal:
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    if isinstance(v, str):
        cleaned = v.replace(",", "").strip()
        return Decimal(cleaned)
    raise ValueError(f"unparseable amount: {v!r}")


def _to_int(v: Any) -> int | None:
    if v in (None, ""):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return int(s)
        except ValueError:
            return None
    return None


def parse_workbook(file_bytes: bytes) -> ParseResult:
    from app.transactions.schemas import TransactionIn

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=False, data_only=True
        )
    except Exception as e:
        raise ParseError("WORKBOOK_LOAD_FAILED", reason=str(e)) from e

    sheet_name = _find_target_sheet(wb)
    ws = wb[sheet_name]
    header_row, col_map = _find_header_row(ws)

    # First pass: collect raw date strings for year inference
    date_col = col_map[_norm_header("이용일자")]
    raw_dates: list[str] = []
    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        val = ws.cell(row=row_idx, column=date_col).value
        if val is not None:
            raw_dates.append(str(val))
    year = _infer_year(raw_dates)

    transactions: list[TransactionIn] = []
    errors: list[dict[str, Any]] = []
    data_rows = 0

    merchant_col = col_map[_norm_header("이용가맹점(은행)명")]
    sales_col = col_map.get(_norm_header("매출구분"))
    amount_col = col_map.get(_norm_header("이용금액(해외현지/체크카드)"))
    card_col = col_map.get(_norm_header("이용카드"))
    install_col = col_map.get(_norm_header("할부개월"))

    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        date_val = ws.cell(row=row_idx, column=date_col).value
        if date_val is None:
            continue
        sales_val = ws.cell(row=row_idx, column=sales_col).value if sales_col else None
        if isinstance(sales_val, str) and sales_val.strip() in _SKIP_SALES_TYPES:
            continue
        merchant_val = ws.cell(row=row_idx, column=merchant_col).value
        if not merchant_val:
            continue

        data_rows += 1

        try:
            txn_date = _parse_md_date(date_val, year)
            amount_raw = ws.cell(row=row_idx, column=amount_col).value if amount_col else None
            amount = _to_decimal(amount_raw) if amount_raw is not None else Decimal("0")

            sales_str = str(sales_val).strip() if sales_val else ""
            is_canceled = sales_str == "취소" or amount < Decimal("0")

            card_val = ws.cell(row=row_idx, column=card_col).value if card_col else None
            last4 = str(card_val).strip() if card_val is not None else None
            if last4 and not (len(last4) == 4 and last4.isdigit()):
                # 정상이 아니면 무시
                last4 = None

            install_raw = ws.cell(row=row_idx, column=install_col).value if install_col else None
            installment = _to_int(install_raw)

            raw_row = {}
            for header, c_idx in col_map.items():
                raw_row[header] = ws.cell(row=row_idx, column=c_idx).value

            transactions.append(TransactionIn(
                txn_date=txn_date,
                txn_time=None,  # 우리카드는 시간 정보 없음
                amount=amount.copy_abs() if is_canceled else amount,
                merchant_raw=str(merchant_val).strip(),
                approval_no=None,  # 우리카드 명세서에 승인번호 없음
                card_last4=last4,
                installment_months=installment,
                is_canceled=is_canceled,
                raw_row=raw_row,
            ))
        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})

    if data_rows == 0:
        raise ParseError("EMPTY_SHEET", sheet=sheet_name)

    return ParseResult(rows_total=data_rows, transactions=transactions, parse_errors=errors)
